import tkinter.messagebox as messagebox

def exportar_a_excel(datos, nombre_archivo, encabezados):
    """Exporta datos a un archivo Excel usando openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = nombre_archivo.replace('.xlsx', '')

        for col_num, encabezado in enumerate(encabezados, start=1):
            cell = ws.cell(row=1, column=col_num, value=encabezado)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        for row_num, fila in enumerate(datos, start=2):
            for col_num, valor in enumerate(fila, start=1):
                ws.cell(row=row_num, column=col_num, value=str(valor))

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(nombre_archivo)
        messagebox.showinfo("Éxito", f"Datos exportados a {nombre_archivo}")
    except ImportError:
        messagebox.showerror("Error", "Instala 'openpyxl': pip install openpyxl")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar a Excel: {e}")

def exportar_a_pdf(datos, nombre_archivo, encabezados, titulo_tabla="Datos"):
    """Exporta datos a un archivo PDF usando reportlab"""
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch

        doc = SimpleDocTemplate(nombre_archivo, pagesize=A4)
        elements = []

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,
        )
        title = Paragraph(titulo_tabla, title_style)
        elements.append(title)

        elements.append(Spacer(1, 12))

        data_for_table = [encabezados] + datos
        table = Table(data_for_table)

        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)
        doc.build(elements)
        messagebox.showinfo("Éxito", f"Datos exportados a {nombre_archivo}")
    except ImportError:
        messagebox.showerror("Error", "Instala 'reportlab': pip install reportlab")
    except Exception as e:
        messagebox.showerror("Error", f"Error al exportar a PDF: {e}")